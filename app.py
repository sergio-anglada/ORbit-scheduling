# Para ejecutar en local: python -m streamlit run app.py

import streamlit as st
import time
import io
from ortools.sat.python import cp_model

st.set_page_config(page_title="ORbit - Scheduling", layout="wide")

st.title("ORbit — Scheduling")
st.markdown("*Optimización matemática aplicada a la toma de decisiones empresarial*")

st.markdown("""
Soy **Sergio Anglada**, investigador en el Departamento de Métodos Estadísticos de la Universidad de Zaragoza. 
Esta herramienta nació de un problema real: los conserjes de los edificios de Matemáticas y Geológicas de la 
Facultad de Ciencias nos pidieron ayuda para organizar sus turnos semanales. Junto con mis compañeros 
**Herminia I. Calvete, Carmen Galé, Aitor Hernández y José A. Iranzo**, lo resolvimos con un modelo de 
programación entera y publicamos el resultado en la revista **conCIENCIAS** (mayo 2025).

A raíz de ese trabajo, nació **ORbit**: una plataforma de optimización matemática para empresas que toman 
decisiones complejas a mano, con Excel, o directamente a ojo. El objetivo es claro: ahorrar tiempo, dinero 
y trabajo mediante modelos que garantizan la solución óptima. La planificación de turnos es el primer problema. 
Vendrán más.

*Esta versión es un primer prototipo funcional de un módulo de scheduling.*
""")

st.divider()

# ─── Parámetros en sidebar ───────────────────────────────────────────────────
with st.sidebar:
    st.header("Parámetros")
    n_trabajadores = st.number_input("Número de trabajadores", 2, 20, 6)
    n_edificios    = st.number_input("Número de edificios",    1, 10, 2)
    n_semanas      = st.slider("Semanas a planificar", 1, 30, 15)
    dif_edificios  = st.slider("Diferencia máxima entre edificios (Δ)", 0, 10, 2)
    timeout        = st.slider("Tiempo máximo de resolución (s)", 10, 1800, 300)

    st.divider()
    personalizar_cob = st.checkbox("Personalizar cobertura por edificio")

    cob_manana_e = {}
    cob_tarde_e  = {}

    if personalizar_cob:
        for e in range(int(n_edificios)):
            st.markdown(f"**Edificio {e+1}**")
            cob_manana_e[e] = st.number_input(f"Mañana", 1, 5, 2, key=f"cob_m_{e}")
            cob_tarde_e[e]  = st.number_input(f"Tarde",  1, 5, 1, key=f"cob_t_{e}")
    else:
        cob_manana_base = st.number_input("Conserjes mañana por edificio", 1, 5, 2)
        cob_tarde_base  = st.number_input("Conserjes tarde por edificio",  1, 5, 1)
        for e in range(int(n_edificios)):
            cob_manana_e[e] = cob_manana_base
            cob_tarde_e[e]  = cob_tarde_base

    st.divider()
    personalizar = st.checkbox("Personalizar nombres")

    nombres_trabajadores = []
    nombres_edificios_input = []

    if personalizar:
        st.markdown("**Trabajadores**")
        for i in range(int(n_trabajadores)):
            nombre = st.text_input(f"Trabajador {i+1}", value=f"Trabajador {i+1}", key=f"trab_{i}")
            nombres_trabajadores.append(nombre)
        st.markdown("**Edificios**")
        for e in range(int(n_edificios)):
            nombre = st.text_input(f"Edificio {e+1}", value=f"Edificio {e+1}", key=f"edif_{e}")
            nombres_edificios_input.append(nombre)
    else:
        nombres_trabajadores = [f"Trabajador {i+1}" for i in range(int(n_trabajadores))]
        nombres_edificios_input = [f"Edificio {e+1}" for e in range(int(n_edificios))]

    st.divider()
    st.markdown("**Restricciones opcionales**")
    usar_r4 = st.checkbox("Máx. 3 semanas consecutivas en mismo edificio", value=True)
    usar_r5 = st.checkbox("Tras tarde, obligatorio mañana las 2 semanas siguientes", value=True)
    usar_r6 = st.checkbox("Rotación de tarde entre edificios cada 3 semanas", value=True)
    usar_r9 = st.checkbox("Equilibrio de semanas entre edificios", value=True)

resolver = st.button("Resolver", type="primary", use_container_width=True)

# ─── Solver ──────────────────────────────────────────────────────────────────
if resolver:

    # ── Validación previa ────────────────────────────────────────────────
    demanda_total = sum(cob_manana_e[e] + cob_tarde_e[e] for e in range(int(n_edificios)))
    if n_trabajadores < demanda_total:
        st.error(
            f"❌ No hay suficientes trabajadores para cubrir todos los turnos. "
            f"Con la cobertura definida se necesitan exactamente **{demanda_total} trabajadores**, "
            f"pero solo hay **{n_trabajadores}**."
        )
        st.stop()
    elif n_trabajadores > demanda_total:
        st.error(
            f"❌ Hay más trabajadores que huecos disponibles. "
            f"Con la cobertura definida se necesitan exactamente **{demanda_total} trabajadores**, "
            f"pero hay **{n_trabajadores}**."
        )
        st.stop()

    I = range(n_trabajadores)
    E = range(n_edificios)
    K = range(2)
    T = range(n_semanas)

    nombres_edificios = nombres_edificios_input

    model  = cp_model.CpModel()
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = timeout
    solver.parameters.log_search_progress = False

    # Variables principales
    x = {}
    for i in I:
        for e in E:
            for k in K:
                for t in T:
                    x[i,e,k,t] = model.new_bool_var(f"x_{i}_{e}_{k}_{t}")

    # Variables auxiliares: coincidencia entre trabajadores
    y = {}
    for i in I:
        for j in I:
            if i != j:
                for e in E:
                    for t in T:
                        y[i,j,e,t] = model.new_bool_var(f"y_{i}_{j}_{e}_{t}")

    # z_max: máximo de semanas que coinciden dos trabajadores
    z_max = model.new_int_var(0, n_semanas * n_edificios, "z_max")

    # ── Restricciones ──────────────────────────────────────────────────────

    # R1: Cada trabajador hace exactamente 1 turno por semana (en algún edificio)
    for i in I:
        for t in T:
            model.add(sum(x[i,e,k,t] for e in E for k in K) == 1)

    # R2: Cobertura mínima mañana por edificio
    for e in E:
        for t in T:
            model.add(sum(x[i,e,0,t] for i in I) == cob_manana_e[e])

    # R3: Cobertura mínima tarde por edificio
    for e in E:
        for t in T:
            model.add(sum(x[i,e,1,t] for i in I) == cob_tarde_e[e])

    # R4: Máximo 3 semanas en el mismo edificio en ventana de 4
    if usar_r4:
        for i in I:
            for e in E:
                for t in range(n_semanas - 3):
                    model.add(sum(x[i,e,k,t+d] for k in K for d in range(4)) <= 3)

    # R5: Tras turno de tarde, siguiente turno debe ser de mañana
    if usar_r5:
        for i in I:
            for t in range(n_semanas - 2):
                model.add(
                    sum(x[i,e,1,t] for e in E) + 1 <=
                    sum(x[i,e,0,t+1] + x[i,e,0,t+2] for e in E)
                )

    # R6: Rotación de tarde entre edificios (si n_edificios >= 2)
    if usar_r6 and n_edificios >= 2:
        for i in I:
            for t in range(n_semanas - 3):
                model.add(x[i,0,1,t] == x[i,1,1,t+3])
                model.add(x[i,1,1,t] == x[i,0,1,t+3])

    # R7: Definición de y (coincidencia en mañana)
    for i in I:
        for j in I:
            if i != j:
                for e in E:
                    for t in T:
                        model.add(y[i,j,e,t] <= x[i,e,0,t])
                        model.add(y[i,j,e,t] <= x[j,e,0,t])
                        model.add(y[i,j,e,t] >= x[i,e,0,t] + x[j,e,0,t] - 1)

    # R8: z_max = max coincidencias entre cualquier par
    for i in I:
        for j in I:
            if i != j:
                model.add(sum(y[i,j,e,t] for e in E for t in T) <= z_max)

    # R9: Equilibrio entre edificios por trabajador (todos los pares)
    d_plus  = {}
    d_minus = {}
    if usar_r9 and n_edificios >= 2:
        for i in I:
            for e1 in E:
                for e2 in E:
                    if e1 < e2:
                        d_plus[i,e1,e2]  = model.new_int_var(0, n_semanas, f"dp_{i}_{e1}_{e2}")
                        d_minus[i,e1,e2] = model.new_int_var(0, n_semanas, f"dm_{i}_{e1}_{e2}")
                        model.add(
                            sum(x[i,e1,k,t] for k in K for t in T) -
                            sum(x[i,e2,k,t] for k in K for t in T) ==
                            d_plus[i,e1,e2] - d_minus[i,e1,e2]
                        )
                        model.add(d_plus[i,e1,e2]  <= dif_edificios)
                        model.add(d_minus[i,e1,e2] <= dif_edificios)

    model.minimize(z_max)

    # ── Resolver ────────────────────────────────────────────────────────
    progress = st.progress(0, text="Resolviendo...")
    t0 = time.time()

    import threading
    resultado_container = [None]

    def run_solver():
        resultado_container[0] = solver.solve(model)

    thread = threading.Thread(target=run_solver)
    thread.start()

    while thread.is_alive():
        elapsed = time.time() - t0
        pct = min(int(elapsed / timeout * 90), 90)
        progress.progress(pct, text=f"Resolviendo... {elapsed:.0f}s")
        time.sleep(0.5)

    thread.join()
    progress.progress(100, text="Listo")
    elapsed_total = time.time() - t0

    status = resultado_container[0]
    status_name = solver.status_name(status)

    st.session_state["status"] = status
    st.session_state["status_name"] = status_name
    st.session_state["elapsed_total"] = elapsed_total
    st.session_state["solver"] = solver
    st.session_state["vals_x"] = {(i,e,k,t): solver.value(x[i,e,k,t]) for i in I for e in E for k in K for t in T}
    st.session_state["vals_y"] = {(i,j,e,t): solver.value(y[i,j,e,t]) for i in I for j in I if i != j for e in E for t in T}
    st.session_state["obj_val"] = solver.objective_value
    st.session_state["params"] = (list(I), list(E), list(K), list(T), nombres_trabajadores, nombres_edificios, cob_manana_e, cob_tarde_e, dif_edificios, usar_r4, usar_r5, usar_r6, usar_r9)

    st.divider()

    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:

        obj_val = st.session_state["obj_val"]
        es_optimo = status == cp_model.OPTIMAL
        vals_x = st.session_state["vals_x"]
        vals_y = st.session_state["vals_y"]
        I, E, K, T, nombres_trabajadores, nombres_edificios, cob_manana_e, cob_tarde_e, dif_edificios, usar_r4, usar_r5, usar_r6, usar_r9 = st.session_state["params"]
        I, E, K, T = range(len(I)), range(len(E)), range(len(K)), range(len(T))

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Estado", "Óptimo" if es_optimo else "Factible")
        col2.metric("Máx. coincidencias", int(obj_val))
        col3.metric("Tiempo", f"{elapsed_total:.1f}s")
        col4.metric("Semanas", n_semanas)

        if not es_optimo:
            st.warning("Solución factible (no óptima) — se alcanzó el límite de tiempo.")


        # ── Resumen por trabajador ──────────────────────────────────
        import pandas as pd
        st.subheader("Resumen por trabajador")
        rows = []
        for i in I:
            sem_edificios = {e: sum(vals_x[i,e,k,t] for k in K for t in T) for e in E}
            rows.append({
                "Trabajador": nombres_trabajadores[i],
                **{nombres_edificios[e]: int(sem_edificios[e]) for e in E},
                "Total semanas": int(sum(sem_edificios.values()))
            })
        df_resumen = pd.DataFrame(rows)
        st.dataframe(df_resumen, use_container_width=True, hide_index=True)

        # ── Coincidencias entre trabajadores ───────────────────────
        st.subheader("Semanas que coinciden en mañana")

        coin_rows = []
        for i in I:
            for j in I:
                if i < j:
                    total = sum(vals_y[i,j,e,t] for e in E for t in T)
                    coin_rows.append({
                        "Par": f"{nombres_trabajadores[i]} — {nombres_trabajadores[j]}",
                        "Semanas coincidiendo": int(total),
                        "¿Es el máximo?": "⚠️ Sí" if int(total) == int(obj_val) else ""
                    })

        df_coin = pd.DataFrame(coin_rows).sort_values("Semanas coincidiendo", ascending=False)
        st.dataframe(df_coin, use_container_width=True, hide_index=True)

        # ── Cuadrante visual ───────────────────────────────────────
        st.subheader("Cuadrante completo")

        COLORES = [
            "#4e79a7","#f28e2b","#e15759","#76b7b2","#59a14f",
            "#edc948","#b07aa1","#ff9da7","#9c755f","#bab0ac",
            "#1f77b4","#ff7f0e","#2ca02c","#d62728","#9467bd",
            "#8c564b","#e377c2","#7f7f7f","#bcbd22","#17becf",
        ]

        def color_texto(hex_bg):
            h = hex_bg.lstrip("#")
            r, g, b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
            luminancia = (0.299*r + 0.587*g + 0.114*b) / 255
            return "#000000" if luminancia > 0.5 else "#ffffff"

        def celda(trabajadores):
            if not trabajadores:
                return "<td style='padding:4px 8px;'></td>"
            partes = []
            for i in trabajadores:
                bg = COLORES[i % len(COLORES)]
                fg = color_texto(bg)
                partes.append(
                    f"<span style='background:{bg};color:{fg};"
                    f"border-radius:4px;padding:2px 6px;"
                    f"margin:1px;display:inline-block;font-size:0.85em;'>"
                    f"{nombres_trabajadores[i]}</span>"
                )
            return "<td style='padding:4px 8px;'>" + " ".join(partes) + "</td>"

        cabecera_cols = "<th style='padding:4px 8px;'>Semana</th><th style='padding:4px 8px;'>Turno</th>"
        for e in E:
            cabecera_cols += f"<th style='padding:4px 8px;'>{nombres_edificios[e]}</th>"

        filas_html = ""
        for t in T:
            bg_fila_m = "#f0f4f8" if t % 2 == 0 else "#ffffff"
            bg_fila_t = "#dce8f0" if t % 2 == 0 else "#eef4f8"

            fila_m = f"<tr style='background:{bg_fila_m};'>"
            fila_m += f"<td rowspan='2' style='padding:4px 8px;font-weight:bold;vertical-align:middle;'>S{t+1}</td>"
            fila_m += f"<td style='padding:4px 8px;font-size:0.8em;color:#555;'>🌅 Mañana</td>"
            for e in E:
                trabajadores_celda = [i for i in I if vals_x[i,e,0,t] == 1]
                fila_m += celda(trabajadores_celda)
            fila_m += "</tr>"

            fila_t = f"<tr style='background:{bg_fila_t};'>"
            fila_t += f"<td style='padding:4px 8px;font-size:0.8em;color:#555;'>🌆 Tarde</td>"
            for e in E:
                trabajadores_celda = [i for i in I if vals_x[i,e,1,t] == 1]
                fila_t += celda(trabajadores_celda)
            fila_t += "</tr>"

            filas_html += fila_m + fila_t

        leyenda = "<div style='margin-top:12px;display:flex;flex-wrap:wrap;gap:8px;'>"
        for i in I:
            bg = COLORES[i % len(COLORES)]
            fg = color_texto(bg)
            leyenda += (
                f"<span style='background:{bg};color:{fg};"
                f"border-radius:4px;padding:3px 10px;font-size:0.85em;'>"
                f"{nombres_trabajadores[i]}</span>"
            )
        leyenda += "</div>"

        tabla_html = f"""
        <div style='overflow-x:auto;'>
        <table style='border-collapse:collapse;width:100%;font-size:0.9em;'>
          <thead>
            <tr style='background:#2c3e50;color:white;'>
              {cabecera_cols}
            </tr>
          </thead>
          <tbody>
            {filas_html}
          </tbody>
        </table>
        </div>
        {leyenda}
        """
        st.markdown(tabla_html, unsafe_allow_html=True)

        # ── Exportar ───────────────────────────────────────────────
        st.subheader("Exportar")

        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        COLORES_HEX = [
            "4e79a7","f28e2b","e15759","76b7b2","59a14f",
            "edc948","b07aa1","ff9da7","9c755f","bab0ac",
            "1f77b4","ff7f0e","2ca02c","d62728","9467bd",
            "8c564b","e377c2","7f7f7f","bcbd22","17becf",
        ]

        def color_texto_excel(hex_bg):
            r, g, b = int(hex_bg[0:2],16), int(hex_bg[2:4],16), int(hex_bg[4:6],16)
            luminancia = (0.299*r + 0.587*g + 0.114*b) / 255
            return "000000" if luminancia > 0.5 else "ffffff"

        # Mapa trabajador -> índice para colores
        color_map = {nombres_trabajadores[i]: i for i in I}

        # Construir cuadrante con una columna por puesto por edificio
        max_puestos = max(max(cob_manana_e.values()), max(cob_tarde_e.values()))
        cabeceras_cuadrante = ["Semana", "Turno"]
        for e in E:
            for p in range(max_puestos):
                cabeceras_cuadrante.append(f"{nombres_edificios[e]} - Puesto {p+1}")

        cuadrante_rows = []
        for t in T:
            for k in K:
                turno_label = "Mañana" if k == 0 else "Tarde"
                fila = {"Semana": f"S{t+1}", "Turno": turno_label}
                for e in E:
                    trabajadores_celda = [nombres_trabajadores[i] for i in I if vals_x[i,e,k,t] == 1]
                    for p in range(max_puestos):
                        col = f"{nombres_edificios[e]} - Puesto {p+1}"
                        fila[col] = trabajadores_celda[p] if p < len(trabajadores_celda) else ""
                cuadrante_rows.append(fila)
        df_cuadrante = pd.DataFrame(cuadrante_rows, columns=cabeceras_cuadrante)

        # Coincidencias con columnas separadas
        coin_rows_excel = []
        for i in I:
            for j in I:
                if i < j:
                    total = sum(vals_y[i,j,e,t] for e in E for t in T)
                    coin_rows_excel.append({
                        "Trabajador A": nombres_trabajadores[i],
                        "Trabajador B": nombres_trabajadores[j],
                        "Semanas coincidiendo": int(total),
                        "¿Es el máximo?": "Sí" if int(total) == int(obj_val) else ""
                    })
        df_coin_excel = pd.DataFrame(coin_rows_excel).sort_values("Semanas coincidiendo", ascending=False)

        # Generar Excel
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df_cuadrante.to_excel(writer, sheet_name="Cuadrante", index=False)
            df_resumen.to_excel(writer, sheet_name="Resumen por trabajador", index=False)
            df_coin_excel.to_excel(writer, sheet_name="Coincidencias", index=False)

            # ── Formato hoja Cuadrante ──────────────────────────────
            ws = writer.sheets["Cuadrante"]

            # Ajustar ancho de columnas
            for col_idx, col in enumerate(df_cuadrante.columns, 1):
                max_len = max(len(str(col)), df_cuadrante[col].astype(str).map(len).max())
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

            # Colorear celdas de trabajadores
            for row_idx, row in enumerate(df_cuadrante.itertuples(index=False), 2):
                for col_idx, col in enumerate(df_cuadrante.columns, 1):
                    if col in ("Semana", "Turno"):
                        continue
                    valor = getattr(row, row._fields[col_idx - 1])
                    if valor and valor in color_map:
                        idx_trabajador = color_map[valor]
                        hex_bg = COLORES_HEX[idx_trabajador % len(COLORES_HEX)]
                        hex_fg = color_texto_excel(hex_bg)
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                            start_color=hex_bg, end_color=hex_bg, fill_type="solid"
                        )
                        ws.cell(row=row_idx, column=col_idx).font = Font(color=hex_fg, bold=True)

            # ── Formato hoja Resumen por trabajador ────────────────
            ws_res = writer.sheets["Resumen por trabajador"]
            for col_idx, col in enumerate(df_resumen.columns, 1):
                max_len = max(len(str(col)), df_resumen[col].astype(str).map(len).max())
                ws_res.column_dimensions[get_column_letter(col_idx)].width = max_len + 4
            for row_idx, row in enumerate(df_resumen.itertuples(index=False), 2):
                nombre = row[0]
                if nombre in color_map:
                    idx_trabajador = color_map[nombre]
                    hex_bg = COLORES_HEX[idx_trabajador % len(COLORES_HEX)]
                    hex_fg = color_texto_excel(hex_bg)
                    ws_res.cell(row=row_idx, column=1).fill = PatternFill(
                        start_color=hex_bg, end_color=hex_bg, fill_type="solid"
                    )
                    ws_res.cell(row=row_idx, column=1).font = Font(color=hex_fg, bold=True)

            # ── Formato hoja Coincidencias ──────────────────────────
            ws_coin = writer.sheets["Coincidencias"]
            for col_idx, col in enumerate(df_coin_excel.columns, 1):
                max_len = max(len(str(col)), df_coin_excel[col].astype(str).map(len).max())
                ws_coin.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

            # Colorear columnas Trabajador A y Trabajador B
            for row_idx, row in enumerate(df_coin_excel.itertuples(index=False), 2):
                for col_idx, campo in [(1, "Trabajador A"), (2, "Trabajador B")]:
                    valor = getattr(row, row._fields[col_idx - 1])
                    if valor in color_map:
                        idx_trabajador = color_map[valor]
                        hex_bg = COLORES_HEX[idx_trabajador % len(COLORES_HEX)]
                        hex_fg = color_texto_excel(hex_bg)
                        ws_coin.cell(row=row_idx, column=col_idx).fill = PatternFill(
                            start_color=hex_bg, end_color=hex_bg, fill_type="solid"
                        )
                        ws_coin.cell(row=row_idx, column=col_idx).font = Font(color=hex_fg, bold=True)

        buffer.seek(0)

        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                "📥 Descargar Excel",
                buffer,
                "turnos.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_excel"
            )
        with col2:
            csv = df_coin.to_csv(index=False)
            st.download_button(
                "📥 Descargar coincidencias CSV",
                csv,
                "coincidencias.csv",
                "text/csv",
                key="download_csv"
            )

    elif status == cp_model.INFEASIBLE:
        st.error(
            "❌ El problema no tiene solución con las restricciones y parámetros actuales. "
            "Posibles causas: la cobertura definida es incompatible con las restricciones de rotación activas, "
            "o no hay suficientes trabajadores para cumplir todas las condiciones a la vez. "
            "Prueba a desactivar alguna restricción opcional o ajustar la cobertura."
        )
    else:
        st.warning(f"Estado: {status_name}. Prueba a aumentar el tiempo máximo.")
